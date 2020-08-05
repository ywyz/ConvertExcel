"""
寻找表格最大行、列
"""
import openpyxl


class ProcessingForm:
    def __init__(self, sheet):
        self.sheet = sheet

    def GetMaxROW(self):
        """获取今日表格最大行"""
        return self.sheet.max_row

    def GetMaxColumn(self):
        """获取今日表格最大列"""
        return self.sheet.max_column

    def DeleteColumn(self, start, num):
        self.sheet.delete_cols(start, num)

    def DeleteNoneLines(self):
        """删除空行"""
        empty = []
        rows = 1
        for cell in self.sheet['B']:
            if cell.value is not None:
                rows += 1
                continue
            else:
                empty.append(rows)
                rows += 1
                continue

        for x in reversed(empty):
            self.sheet.delete_rows(x)

    def DeleteRedundantData(self):
        """删除多次填写的数据"""

        names = []
        rows = 1
        delete_rows = []
        num = 0
        for cell in self.sheet['D']:
            if cell.value in names:
                delete_rows.append(rows)
                rows += 1

            else:
                names.append(cell.value)
                rows += 1

        for x in reversed(delete_rows):
            self.sheet.delete_rows(x, 1)
            num += 1

    def CompareStudentNumber(self):
        """
        比较各班学生数量
        :return:
        """
        small_one = 0
        small_two = 0
        middle_one = 0
        middle_two = 0
        big_one = 0
        big_two = 0
        rows = 1
        for cell in self.sheet['H']:
            if cell.value == '大班':
                if self.sheet.cell(rows, 9).value == '一班':
                    big_one += 1
                    rows += 1
                    continue
                elif self.sheet.cell(rows, 9).value == '二班':
                    big_two += 1
                    rows += 1
                    continue
            elif cell.value == '中班':
                if self.sheet.cell(rows, 9).value == '一班':
                    middle_one += 1
                    rows += 1
                    continue
                elif self.sheet.cell(rows, 9).value == '二班':
                    middle_two += 1
                    rows += 1
                    continue
            elif cell.value == '小班':
                if self.sheet.cell(rows, 9).value == '一班':
                    small_one += 1
                    rows += 1
                    continue
                elif self.sheet.cell(rows, 9).value == '二班':
                    small_two += 1
                    rows += 1
                    continue

            else:
                rows += 1
                continue

        print("小一班：", small_one)
        print("小二班：", small_two)
        print("中一班：", middle_one)
        print("中二班：", middle_two)
        print("大一班：", big_one)
        print("大二班：", big_two)

    def FindStudentFilledError(self):
        # 建立各班正确工作簿
        row = 1
        num = 1
        temp = 0
        correct_wb = openpyxl.load_workbook("correct.xlsx")
        student = correct_wb["student"]

        names = []
        correct_names = {}
        # 将表中姓名添加进数组中
        for cell in self.sheet['D']:
            names.append(cell.value)
        # 将各班正确名字添加进字典中
        for cell in student['A']:
            correct_names[cell.value] = student.cell(num, 2).value
            num += 1

        # 将正确字典中的班级复制进表中
        for cell in self.sheet['D']:
            if cell.value in correct_names.keys():
                if correct_names[cell.value] == "小一":
                    self.sheet.cell(row, 8).value = '小班'
                    self.sheet.cell(row, 9).value = '一班'
                    row += 1
                    continue
                elif correct_names[cell.value] == "小二":
                    self.sheet.cell(row, 8).value = '小班'
                    self.sheet.cell(row, 9).value = '二班'
                    row += 1
                    continue
                elif correct_names[cell.value] == "中一":
                    self.sheet.cell(row, 8).value = '中班'
                    self.sheet.cell(row, 9).value = '一班'
                    row += 1
                    continue
                elif correct_names[cell.value] == "中二":
                    self.sheet.cell(row, 8).value = '中班'
                    self.sheet.cell(row, 9).value = '二班'
                    row += 1
                    continue
                elif correct_names[cell.value] == "大一":
                    self.sheet.cell(row, 8).value = '大班'
                    self.sheet.cell(row, 9).value = '一班'
                    row += 1
                    continue
                elif correct_names[cell.value] == "大二":
                    self.sheet.cell(row, 8).value = '大班'
                    self.sheet.cell(row, 9).value = '二班'
                    row += 1
                    continue
            elif cell.value == "孩子姓名":
                row += 1
                continue
            else:
                print(cell.value, "不在姓名表中，请核实。")
                row += 1
                continue

            # 检查是否有没填的人

        for key, value in correct_names.items():
            if key in names:
                continue
            else:
                print(key, value, "没填")
                temp += 1
        print("共有", temp, "人没填")

    def CompareTeacherNumber(self):
        num = 0
        for cell in self.sheet['D']:
            if cell.value == "教职工信息—姓名":
                continue

            else:
                num += 1
        print("教职工：", num)

    def FindTeacherFilledError(self):

        correct_wb = openpyxl.load_workbook("correct.xlsx")
        teacher = correct_wb["teacher"]

        names = []
        correct_names = []
        # 将表中姓名添加进数组中
        for cell in self.sheet['D']:
            if cell.value == '教职工信息—姓名':
                continue
            else:
                names.append(cell.value)
        # 将教职工名字添加进数组中
        for cell in teacher['A']:
            correct_names.append(cell.value)
        #
        for name in names:
            if name in correct_names:
                continue
            else:
                print(name, "填错了")

        for correct in correct_names:
            if correct in names:
                continue
            else:
                print(correct, "没填")

    def CopyClass(self, small_one, small_two, middle_one, middle_two, big_one, big_two):
        """分到各班表格中"""
        for n in range(1, self.sheet.max_column + 1):
            cell1 = self.sheet.cell(1, column=n).value
            small_one.cell(1, column=n).value = cell1
            small_two.cell(1, column=n).value = cell1
            middle_one.cell(1, column=n).value = cell1
            middle_two.cell(1, column=n).value = cell1
            big_one.cell(1, column=n).value = cell1
            big_two.cell(1, column=n).value = cell1

        for m in range(2, self.sheet.max_row + 1):  # 遍历行
            for n in range(1, self.sheet.max_column + 1):
                if self.sheet.cell(row=m, column=8).value == "大班":
                    if self.sheet.cell(row=m, column=9).value == "一班":
                        cell1 = self.sheet.cell(row=m, column=n).value
                        big_one.cell(row=m, column=n).value = cell1
                    elif self.sheet.cell(row=m, column=9).value == "二班":
                        cell1 = self.sheet.cell(row=m, column=n).value
                        big_two.cell(row=m, column=n).value = cell1
                if self.sheet.cell(row=m, column=8).value == "中班":
                    if self.sheet.cell(row=m, column=9).value == "一班":
                        cell1 = self.sheet.cell(row=m, column=n).value
                        middle_one.cell(row=m, column=n).value = cell1
                    if self.sheet.cell(row=m, column=9).value == "二班":
                        cell1 = self.sheet.cell(row=m, column=n).value
                        middle_two.cell(row=m, column=n).value = cell1
                if self.sheet.cell(row=m, column=8).value == "小班":
                    if self.sheet.cell(row=m, column=9).value == "一班":
                        cell1 = self.sheet.cell(row=m, column=n).value
                        small_one.cell(row=m, column=n).value = cell1
                    if self.sheet.cell(row=m, column=9).value == "二班":
                        cell1 = self.sheet.cell(row=m, column=n).value
                        small_two.cell(row=m, column=n).value = cell1

    def AddSorted(self):
        self.sheet.auto_filter.ref = "A1:Z50"
        self.sheet.auto_filter.add_sort_condition("J1:J50")

    def ChangeNum(self):
        num = 1
        for cell in self.sheet['A']:
            if cell.value == "序号":
                continue

            else:
                cell.value = num
                num += 1
